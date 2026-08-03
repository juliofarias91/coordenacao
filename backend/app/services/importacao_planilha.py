"""Lê as planilhas de auditoria que a coordenação já preenche à mão.

ISTO É UMA PONTE, NÃO O MODELO DE DADOS. A plataforma tem um caminho próprio
para auditoria — critério, checklist, round, resultado — e nada aqui passa por
ele: o importador lê o arquivo Excel e guarda o que leu em duas tabelas suas
(`importacao_planilha` e `importacao_item`), que só alimentam o dashboard de
importações. Foi feito assim de propósito e sob pressa, para haver número na
tela a partir das planilhas que existem hoje, sem mexer no motor de auditoria.
Quando os dados forem migrados para o caminho real, isto sai inteiro.

O QUE O ARQUIVO REAL ENSINOU (14 planilhas do projeto DANTE 2, lidas antes de
escrever uma linha):

1. **A aba `BASE GERAL` é estável**: cabeçalho na linha 5, 17 itens, sempre nas
   mesmas colunas. A aba de LOD 300 NÃO É: seis das oito têm layout de coluna
   diferente. `VERIFICATION` aparece na coluna 9, 11 ou 12; `INFORMATION` existe
   em seis e falta em duas. Por isso NADA aqui é lido por índice fixo — todo
   acesso passa por `_mapa_de_colunas`, que casa pelo RÓTULO do cabeçalho.

2. **A porcentagem escrita na planilha não é confiável, e há prova.** A aba
   STRC de LOD 300 declara 30% de aprovação; a fórmula dela é
   `=COUNTIF(I6:I33, TRUE)/COUNTA(I6:I65)` — o numerador parou na linha 33 e o
   denominador foi até a 65, porque alguém acrescentou linhas e só arrastou
   metade da conta. O valor certo é 60% (36 aprovados de 60). O arquivo de MECH
   declara exatamente a mesma aprovação que o de FPRT e tem "FPRT-FPRT-DATA" no
   campo de disciplina — outra cópia que ficou pela metade.
   Por isso a aprovação que o dashboard mostra é SEMPRE recontada a partir das
   linhas; a declarada é guardada ao lado, e a tela avisa quando as duas
   divergem. É o argumento inteiro da plataforma numa célula: a planilha não
   erra a auditoria, erra a CONTA sobre ela.

3. **A disciplina sai do NOME DO ARQUIVO, não da célula.** Pela mesma razão: a
   célula está errada em pelo menos dois dos oito arquivos, e o nome do arquivo
   está certo nos catorze.

4. **`FILE NAME` também mente sobre o projeto** — as planilhas do DANTE 2 dizem
   `CPQ04-ARCH-R26`, herança do projeto anterior. Guarda-se o que está escrito,
   sem corrigir: inventar um nome que o arquivo não tem é pior do que mostrar o
   que a coordenação de fato escreveu.

MEDIDO nos catorze arquivos reais em 30/07/2026: 22 auditorias lidas, 833 itens,
nenhuma recusa. Sete delas são a mesma disciplina presente nas duas pastas e são
substituídas pela regra de reimportação, restando 15 no dashboard.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field

# Os códigos de disciplina que aparecem no nome dos arquivos. Lista fechada de
# propósito: sem ela, "DANTE" e "GERAL" (que também são maiúsculas de 5 letras)
# passariam por disciplina.
DISCIPLINAS = {
    "ARCH", "STRC", "ELEC", "MECH", "PLMB", "FPRT", "TCOM", "FALM",
    "CIVL", "LAND", "SITE", "HVAC", "SPKL", "SECU", "INFR", "EQPT", "DEVS",
}

ABA_GERAL = "BASE GERAL"

# Uma linha de LOD 300 sem veredicto é espaçador ou continuação de mesclagem.
_VERDADE = {"TRUE", "VERDADEIRO", "OK", "APPROVED", "APROVADO", "SIM", "YES", "1"}
_FALSIDADE = {"FALSE", "FALSO", "NOT APPROVED", "REPROVADO", "NAO", "NÃO", "NO", "0"}


class PlanilhaInvalida(ValueError):
    """O arquivo não é uma das duas planilhas que este importador conhece."""


@dataclass
class ItemLido:
    ordem: int
    #: O grupo de elemento (LOD 300: FLOOR, CASEWORK…). Vazio na geral, que não
    #: agrupa — os 17 itens dela são uma lista plana.
    grupo: str | None
    item: str
    aprovado: bool
    comentario: str | None
    direcao: str | None


@dataclass
class PlanilhaLida:
    tipo: str  # 'geral' | 'lod300'
    disciplina: str
    modelo: str | None
    versao: str | None
    #: O que a planilha DECLARA, quando declara. Guardado para comparação — não
    #: é o que o dashboard soma.
    aprovacao_declarada: float | None
    itens: list[ItemLido] = field(default_factory=list)

    @property
    def aprovados(self) -> int:
        return sum(1 for i in self.itens if i.aprovado)

    @property
    def aprovacao(self) -> float | None:
        """Recontada a partir das linhas. Ver a nota 2 do cabeçalho."""
        return self.aprovados / len(self.itens) if self.itens else None


# --------------------------------------------------------------------------- util


def _txt(valor: object) -> str:
    """Texto normalizado de uma célula. As planilhas trazem `\\n` e `\\t` dentro
    de rótulo (`'ANALYSED VERSION\\n(V0 - DD/MM/YYYY)'`) e espaço à direita em
    quase todo cabeçalho."""
    if valor is None:
        return ""
    return re.sub(r"\s+", " ", str(valor)).strip()


def _rotulo(valor: object) -> str:
    return _txt(valor).upper()


def _mapa_de_colunas(ws, linha: int, largura: int) -> dict[str, int]:
    """`{RÓTULO: coluna}` da linha de cabeçalho.

    O PRIMEIRO OCUPA. Duas planilhas repetem `LOD` em duas colunas (uma é o
    nível, a outra é `LOI`), e a segunda não acrescenta nada — ficar com a
    primeira evita ter de decidir qual é qual sem informação para isso.
    """
    mapa: dict[str, int] = {}
    for c in range(1, largura + 1):
        rot = _rotulo(ws.cell(linha, c).value)
        if rot and rot not in mapa:
            mapa[rot] = c
    return mapa


def _coluna(mapa: dict[str, int], *candidatos: str) -> int | None:
    """A primeira coluna cujo rótulo casa — exato, depois por prefixo.

    A cadeia de candidatos é o que absorve as seis variações de layout: o nome
    do item mora em `INFORMATION` em seis planilhas e em `REVIT PARAMETER` nas
    duas que não têm aquela coluna.
    """
    for nome in candidatos:
        if nome in mapa:
            return mapa[nome]
    for nome in candidatos:
        for rot, col in mapa.items():
            if rot.startswith(nome):
                return col
    return None


def _veredicto(valor: object) -> bool | None:
    """`True`/`False`/`None`. `None` é "esta linha não é um item" — e é o que
    faz o leitor pular espaçador e linha de mesclagem sem precisar saber onde
    a tabela termina."""
    if isinstance(valor, bool):
        return valor
    rot = _rotulo(valor)
    if not rot:
        return None
    # A ordem importa: "NOT APPROVED" contém "APPROVED".
    if rot in _FALSIDADE or rot.startswith("NOT ") or rot.startswith("NÃO "):
        return False
    if rot in _VERDADE:
        return True
    return None


def _limpo(valor: object) -> str | None:
    """Texto de célula de comentário. O traço solto é como a planilha escreve
    "nada a dizer" — guardá-lo encheria o dashboard de hífens."""
    t = _txt(valor)
    return None if t in ("", "-", "–", "—") else t


def _acha_valor_sob(ws, rotulo_alvo: str, ate_linha: int, largura: int) -> object:
    """O valor logo ABAIXO de um rótulo do bloco de cabeçalho.

    `FILE NAME` está na coluna 2 em todas as planilhas, mas `LOD APPROVAL (%)`
    está na 6 em duas e na 9 em quatro. Procurar o rótulo e descer uma linha é o
    que dispensa saber qual.
    """
    for r in range(1, ate_linha):
        for c in range(1, largura + 1):
            if rotulo_alvo in _rotulo(ws.cell(r, c).value):
                return ws.cell(r + 1, c).value
    return None


def _fracao(valor: object) -> float | None:
    """Porcentagem, se for uma. As planilhas põem a versão (`'V1 - 27/07/2026'`)
    na coluna da aprovação em quatro dos arquivos — daí a checagem de faixa e
    não só de tipo."""
    if isinstance(valor, (int, float)) and not isinstance(valor, bool):
        f = float(valor)
        if f > 1:  # alguém digitou 45 em vez de 0,45
            f = f / 100
        return f if 0 <= f <= 1 else None
    return None


def disciplina_do_arquivo(nome: str) -> str | None:
    """O código de disciplina no nome do arquivo. Ver a nota 3 do cabeçalho."""
    for token in re.findall(r"[A-Za-z]{3,5}", nome.upper()):
        if token in DISCIPLINAS:
            return token
    return None


# ----------------------------------------------------------------- as duas abas


def _ler_geral(ws, disciplina: str) -> PlanilhaLida | None:
    """A aba `BASE GERAL` — os 17 itens da auditoria geral."""
    largura = min(ws.max_column or 1, 20)
    linha = next(
        (r for r in range(1, 16) if _rotulo(ws.cell(r, 1).value) == "INFORMATION"), None
    )
    if linha is None:
        return None

    mapa = _mapa_de_colunas(ws, linha, largura)
    c_item = _coluna(mapa, "INFORMATION")
    c_ver = _coluna(mapa, "VERIFICATION")
    if not c_item or not c_ver:
        return None
    c_com = _coluna(mapa, "COMENTARY", "COMMENTARY", "COMMENTS")
    c_dir = _coluna(mapa, "DIRECTION")

    itens: list[ItemLido] = []
    for r in range(linha + 1, min(ws.max_row, linha + 200) + 1):
        nome = _txt(ws.cell(r, c_item).value)
        veredicto = _veredicto(ws.cell(r, c_ver).value)
        # Linha 18 dos arquivos reais tem nota na coluna I e nada mais: sem nome
        # e sem veredicto, não é item.
        if not nome or veredicto is None:
            continue
        itens.append(
            ItemLido(
                ordem=len(itens),
                grupo=None,
                item=nome,
                aprovado=veredicto,
                comentario=_limpo(ws.cell(r, c_com).value) if c_com else None,
                direcao=_limpo(ws.cell(r, c_dir).value) if c_dir else None,
            )
        )
    if not itens:
        return None

    return PlanilhaLida(
        tipo="geral",
        disciplina=disciplina,
        # Duas linhas acima do cabeçalho, na primeira coluna — estável nas oito.
        modelo=_txt(ws.cell(linha - 2, 1).value) or None,
        versao=None,
        # NA GERAL O VALOR FICA ACIMA DO RÓTULO, não abaixo: "APPROVED (%)" é
        # cabeçalho de coluna (linha 5) e a porcentagem está na linha 2, no topo
        # do bloco. É o oposto do LOD 300, onde rótulo e valor são duas linhas
        # empilhadas — por isso são dois caminhos e não um.
        aprovacao_declarada=_declarada_acima(ws, mapa, linha),
        itens=itens,
    )


def _declarada_acima(ws, mapa: dict[str, int], linha: int) -> float | None:
    col = _coluna(mapa, "APPROVED (%)", "APPROVED")
    if not col:
        return None
    for r in range(1, linha):
        f = _fracao(ws.cell(r, col).value)
        if f is not None:
            return f
    return None


def _ler_lod300(ws, disciplina: str) -> PlanilhaLida | None:
    """A aba de LOD 300 — a de layout instável. Ver a nota 1 do cabeçalho."""
    largura = min(ws.max_column or 1, 20)
    linha = None
    for r in range(1, 16):
        rotulos = {_rotulo(ws.cell(r, c).value) for c in range(1, largura + 1)}
        if "ELEMENT" in rotulos and "VERIFICATION" in rotulos:
            linha = r
            break
    if linha is None:
        return None

    mapa = _mapa_de_colunas(ws, linha, largura)
    c_ver = _coluna(mapa, "VERIFICATION")
    c_grupo = _coluna(mapa, "ELEMENT")
    # A cadeia que absorve as duas planilhas sem coluna INFORMATION.
    c_item = _coluna(mapa, "INFORMATION", "REVIT PARAMETER", "PARAMETER")
    if not c_ver or not c_item:
        return None
    c_com = _coluna(mapa, "COMMENTS", "SUPPLIER COMMENTS", "SUPPLIERS COMMENTS")

    itens: list[ItemLido] = []
    grupo = None
    for r in range(linha + 1, (ws.max_row or linha) + 1):
        # O grupo vem MESCLADO: só a primeira linha de cada bloco o traz, e as
        # seguintes ficam vazias. Carregar o último visto é o que devolve a
        # coluna que a mesclagem apagou.
        if c_grupo:
            novo = _txt(ws.cell(r, c_grupo).value)
            if novo:
                grupo = novo
        veredicto = _veredicto(ws.cell(r, c_ver).value)
        if veredicto is None:
            continue
        nome = _txt(ws.cell(r, c_item).value)
        if not nome:
            continue
        itens.append(
            ItemLido(
                ordem=len(itens),
                grupo=grupo,
                item=nome,
                aprovado=veredicto,
                comentario=_limpo(ws.cell(r, c_com).value) if c_com else None,
                direcao=None,
            )
        )
    if not itens:
        return None

    return PlanilhaLida(
        tipo="lod300",
        disciplina=disciplina,
        modelo=_txt(_acha_valor_sob(ws, "FILE NAME", linha, largura)) or None,
        versao=_txt(_acha_valor_sob(ws, "VERSION", linha, largura)) or None,
        aprovacao_declarada=_fracao(_acha_valor_sob(ws, "APPROVAL", linha, largura)),
        itens=itens,
    )


def ler(nome_arquivo: str, conteudo: bytes) -> list[PlanilhaLida]:
    """Todas as auditorias que houver no arquivo.

    UM ARQUIVO PODE TRAZER AS DUAS. Nos arquivos reais, os da pasta `LOD 300`
    têm `BASE GERAL` E a aba da disciplina — a mesma pasta de trabalho serve às
    duas auditorias. Por isso a função devolve uma LISTA: importar o arquivo de
    ELEC traz a geral e a de LOD 300 numa tacada.
    """
    # `openpyxl` DENTRO da função, como em `services/exports.py`: são ~4 s de
    # import que toda subida do servidor pagaria por uma biblioteca que só entra
    # quando alguém sobe uma planilha.
    import io

    import openpyxl

    try:
        wb = openpyxl.load_workbook(io.BytesIO(conteudo), data_only=True, read_only=False)
    except Exception as e:  # noqa: BLE001 — qualquer falha aqui é "não é xlsx"
        raise PlanilhaInvalida(f"não foi possível abrir como Excel: {e}") from e

    try:
        disciplina_arquivo = disciplina_do_arquivo(nome_arquivo)
        achadas: list[PlanilhaLida] = []

        for aba in wb.sheetnames:
            nome = _txt(aba).upper()
            if "GUIDE" in nome:
                continue
            ws = wb[aba]
            if nome == ABA_GERAL:
                lida = _ler_geral(ws, disciplina_arquivo or "N/D")
            else:
                # A aba de LOD 300 se chama pela disciplina (`ELEC`, ` ARCH`) —
                # é a segunda fonte quando o nome do arquivo não a entrega.
                disc = disciplina_arquivo or (nome if nome in DISCIPLINAS else "N/D")
                lida = _ler_lod300(ws, disc)
            if lida:
                achadas.append(lida)

        if not achadas:
            raise PlanilhaInvalida(
                "nenhuma aba reconhecida: esperava 'BASE GERAL' (auditoria geral) "
                "ou uma aba de LOD 300 com as colunas ELEMENT e VERIFICATION"
            )
        return achadas
    finally:
        wb.close()
