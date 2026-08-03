"""SP-207 · Relatório de RNC (PDF) e controle (Excel).

Ambos são *gerados*, nunca mantidos. O relatório traz só o que reprovou —
é o que a especificação chama de "documento gerado sob demanda" — e sai
bilíngue a partir dos rótulos `nome_pt`/`nome_en` dos critérios.
"""

from __future__ import annotations

import io
import uuid
from datetime import UTC, datetime

# O `openpyxl` NÃO É IMPORTADO AQUI, e sim dentro de `controle_xlsx` — a única
# função que o usa.
#
# Ele custava ~5,9 s no import da aplicação, o maior item isolado do boot, e
# este módulo é alcançado pelo router. O preço era pago em toda subida do
# servidor, inclusive nos reinícios do `--reload`, por uma biblioteca que só
# entra quando alguém baixa a planilha de controle.
#
# O PDF, logo acima, não tem custo equivalente: ele é montado com string, sem
# dependência de terceiro.
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    KeepTogether,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from app.models import (
    Auditoria,
    Criterio,
    Empresa,
    Modelo,
    NaoConformidade,
    Projeto,
    ResultadoCheck,
    VersaoModelo,
)
from app.models.enums import CheckStatus
from app.services.painel import painel_de_controle

AZUL = colors.HexColor("#2547B0")
CINZA = colors.HexColor("#586071")
LINHA = colors.HexColor("#E1E6EE")


def _t(pt: str, en: str, idioma: str) -> str:
    return pt if idioma == "pt" else en


def _escapar(texto: str | None) -> str:
    """ReportLab interpreta o texto como mini-HTML; `&`/`<` precisam escapar."""
    return (
        (texto or "")
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
    )


# --------------------------------------------------------------------- PDF
def relatorio_pdf(db: Session, modelo_id: uuid.UUID, *, idioma: str = "pt") -> bytes:
    """Relatório de auditoria de um modelo: metodologia, resumo e as RNCs.

    Escopo = último round de cada checklist do modelo. É o que a coordenação
    manda para o fornecedor.
    """
    modelo = db.get(Modelo, modelo_id)
    if modelo is None:
        raise ValueError("modelo não encontrado")

    projeto = db.get(Projeto, modelo.projeto_id)
    instaladora = db.get(Empresa, modelo.instaladora_id) if modelo.instaladora_id else None

    versoes = list(
        db.execute(
            select(VersaoModelo)
            .where(VersaoModelo.modelo_id == modelo_id)
            .order_by(VersaoModelo.created_at.desc())
        ).scalars()
    )
    versao = versoes[0] if versoes else None

    auditorias = (
        list(
            db.execute(
                select(Auditoria)
                .where(Auditoria.versao_id == versao.id)
                .order_by(Auditoria.checklist, Auditoria.round.desc().nulls_last())
            ).scalars()
        )
        if versao
        else []
    )
    # Uma auditoria por checklist: a de maior round.
    ultimas: dict[str, Auditoria] = {}
    for a in auditorias:
        ultimas.setdefault(a.checklist.value, a)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=20 * mm,
        rightMargin=20 * mm,
        topMargin=18 * mm,
        bottomMargin=18 * mm,
        title=f"{_t('Relatório de auditoria', 'Audit report', idioma)} — {modelo.codigo}",
        author="SPBIM",
    )

    base = getSampleStyleSheet()
    est_titulo = ParagraphStyle(
        "titulo", parent=base["Title"], fontSize=17, textColor=AZUL, alignment=TA_LEFT
    )
    est_h2 = ParagraphStyle(
        "h2", parent=base["Heading2"], fontSize=12, textColor=AZUL, spaceBefore=14
    )
    est_corpo = ParagraphStyle("corpo", parent=base["BodyText"], fontSize=9, leading=13)
    est_meta = ParagraphStyle("meta", parent=est_corpo, textColor=CINZA, fontSize=8.5)

    fluxo: list = [
        Paragraph(
            "SPBIM",
            ParagraphStyle("marca", parent=base["Normal"], fontSize=9, textColor=CINZA),
        ),
        Paragraph(_t("Relatório de auditoria BIM", "BIM audit report", idioma), est_titulo),
        Paragraph(
            f"{_escapar(projeto.codigo if projeto else '')} · "
            f"<b>{_escapar(modelo.codigo)}</b> · "
            f"{_t('versão', 'version', idioma)} {_escapar(versao.versao if versao else '—')}",
            est_corpo,
        ),
        Paragraph(
            f"{_t('Projetista', 'Designer', idioma)}: "
            f"{_escapar(instaladora.nome if instaladora else '—')} · "
            f"{_t('Emitido em', 'Issued on', idioma)}: "
            f"{datetime.now(UTC).strftime('%d/%m/%Y')}",
            est_meta,
        ),
        Spacer(1, 8 * mm),
    ]

    # --- metodologia (seção fixa) -----------------------------------------
    fluxo += [
        Paragraph(_t("1. Metodologia", "1. Methodology", idioma), est_h2),
        Paragraph(
            _t(
                "A auditoria verifica o modelo entregue contra os critérios da biblioteca do "
                "projeto, derivados do PEB e do A5.37. Cada critério recebe um dos estados "
                "aprovado, reprovado, pendente ou não aplicável. O percentual de aprovação "
                "considera apenas os critérios aplicáveis — os marcados como N/A saem do "
                "denominador. Somente os itens reprovados geram não-conformidade.",
                "The audit checks the delivered model against the project criteria library, "
                "derived from the BEP and A5.37. Each criterion gets one of approved, rejected, "
                "pending or not applicable. The approval percentage considers only applicable "
                "criteria — those marked N/A leave the denominator. Only rejected items raise a "
                "non-conformity.",
                idioma,
            ),
            est_corpo,
        ),
    ]

    # --- resumo por checklist ---------------------------------------------
    fluxo += [Paragraph(_t("2. Resumo", "2. Summary", idioma), est_h2)]

    cabecalho = [
        _t("Auditoria", "Audit", idioma),
        _t("Round", "Round", idioma),
        _t("Estado", "State", idioma),
        _t("Aprovação", "Approval", idioma),
        _t("Reprovados", "Rejected", idioma),
    ]
    linhas_resumo = [cabecalho]
    for checklist, auditoria in ultimas.items():
        reprovados = db.execute(
            select(ResultadoCheck).where(
                ResultadoCheck.auditoria_id == auditoria.id,
                ResultadoCheck.status == CheckStatus.REPROVADO,
            )
        ).scalars()
        linhas_resumo.append(
            [
                checklist.upper(),
                str(auditoria.round or "—"),
                auditoria.estado.value,
                f"{auditoria.aprovacao_pct:.0f}%" if auditoria.aprovacao_pct is not None else "—",
                str(len(list(reprovados))),
            ]
        )

    if len(linhas_resumo) == 1:
        fluxo.append(
            Paragraph(
                _t("Sem auditoria nesta versão.", "No audit on this version.", idioma),
                est_corpo,
            )
        )
    else:
        tabela = Table(linhas_resumo, colWidths=[45 * mm, 18 * mm, 32 * mm, 28 * mm, 27 * mm])
        tabela.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F6F8FB")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), CINZA),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8.5),
                    ("GRID", (0, 0), (-1, -1), 0.4, LINHA),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ]
            )
        )
        fluxo.append(tabela)

    # --- não-conformidades -------------------------------------------------
    ncs = list(
        db.execute(
            select(NaoConformidade)
            .options(selectinload(NaoConformidade.comentarios))
            .where(
                NaoConformidade.auditoria_id.in_(
                    [a.id for a in ultimas.values()] or [uuid.uuid4()]
                )
            )
            .order_by(NaoConformidade.created_at)
        ).scalars()
    )

    fluxo += [
        PageBreak(),
        Paragraph(
            _t("3. Não-conformidades", "3. Non-conformities", idioma)
            + f" ({len(ncs)})",
            est_h2,
        ),
    ]

    if not ncs:
        fluxo.append(
            Paragraph(
                _t(
                    "Nenhuma não-conformidade registrada neste round.",
                    "No non-conformity recorded in this round.",
                    idioma,
                ),
                est_corpo,
            )
        )

    for i, nc in enumerate(ncs, start=1):
        criterio = db.get(Criterio, nc.criterio_id) if nc.criterio_id else None
        responsavel = db.get(Empresa, nc.responsavel_id) if nc.responsavel_id else None
        rotulo = (
            (criterio.nome_pt if idioma == "pt" else criterio.nome_en) if criterio else "—"
        )

        bloco: list = [
            Paragraph(
                f"<b>RNC-{i:03d}</b> · {_escapar(rotulo)}",
                ParagraphStyle("nc", parent=est_corpo, fontSize=10, spaceBefore=10),
            )
        ]
        campos = [
            (_t("Descrição", "Description", idioma), nc.descricao),
            (_t("Recomendação", "Recommendation", idioma), nc.recomendacao),
            (_t("Elementos", "Elements", idioma), nc.elementos),
            (_t("Responsável", "Responsible", idioma), responsavel.nome if responsavel else None),
            (_t("Prazo", "Due date", idioma), nc.prazo.strftime("%d/%m/%Y") if nc.prazo else None),
            (_t("Situação", "Status", idioma), nc.status),
        ]
        for rotulo_campo, valor in campos:
            if valor:
                bloco.append(
                    Paragraph(f"<b>{rotulo_campo}:</b> {_escapar(str(valor))}", est_corpo)
                )

        for comentario in nc.comentarios:
            bloco.append(
                Paragraph(
                    f"<i>{_t('Fornecedor', 'Supplier', idioma)}:</i> "
                    f"{_escapar(comentario.texto)}",
                    est_meta,
                )
            )

        fluxo.append(KeepTogether(bloco))

    doc.build(fluxo)
    return buffer.getvalue()


# ------------------------------------------------------------------- Excel
def controle_xlsx(db: Session, projeto_id: uuid.UUID) -> bytes:
    """Controle modelo × status. Substitui a planilha, gerado do mesmo dado."""
    # Import local — ver a nota no topo do módulo. `openpyxl` é o item mais caro
    # do boot da aplicação e só serve a esta função.
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    projeto = db.get(Projeto, projeto_id)
    if projeto is None:
        raise ValueError("projeto não encontrado")

    linhas = painel_de_controle(db, projeto_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Controle"

    ws["A1"] = f"SPBIM · {projeto.codigo} — {projeto.nome}"
    ws["A1"].font = Font(size=13, bold=True, color="2547B0")
    ws["A2"] = f"Gerado em {datetime.now(UTC).strftime('%d/%m/%Y %H:%M')} UTC"
    ws["A2"].font = Font(size=9, color="586071")

    cabecalho = [
        "Modelo",
        "Disciplina",
        "Macro",
        "Instaladora",
        "Modeladora",
        "Versão",
        "Formato",
        "Round",
        "Estado",
        "Aprovação (%)",
        "NCs abertas",
        "Publicado em",
    ]
    ws.append([])
    ws.append(cabecalho)

    linha_cabecalho = ws.max_row
    fundo = PatternFill("solid", fgColor="F6F8FB")
    for coluna in range(1, len(cabecalho) + 1):
        celula = ws.cell(row=linha_cabecalho, column=coluna)
        celula.font = Font(bold=True, size=9, color="586071")
        celula.fill = fundo
        celula.alignment = Alignment(vertical="center")

    for linha in linhas:
        ws.append(
            [
                linha.codigo,
                linha.disciplina_codigo,
                linha.macro,
                linha.instaladora,
                linha.modeladora,
                linha.versao,
                linha.formato,
                linha.round,
                linha.estado,
                float(linha.aprovacao_pct) if linha.aprovacao_pct is not None else None,
                linha.ncs_abertas,
                linha.publicado_em.strftime("%d/%m/%Y") if linha.publicado_em else None,
            ]
        )

    larguras = [34, 14, 7, 20, 20, 9, 9, 8, 16, 14, 12, 14]
    for i, largura in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(i)].width = largura

    ws.freeze_panes = ws.cell(row=linha_cabecalho + 1, column=1)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
