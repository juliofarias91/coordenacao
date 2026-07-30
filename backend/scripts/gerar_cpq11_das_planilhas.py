"""Gera o YAML do CPQ11 a partir das planilhas de controle reais de `Bases/`.

NÃO INVENTA NADA. Empresas, disciplinas, áreas e nomes de arquivo saem das
planilhas que a coordenação usa hoje:

  BASE (do LOD500)        listas de validação: disciplinas, subsistemas, áreas,
                          empresas com código
  LOD 400 - ADMN          modelo x empresa executante x empresa de modelagem
  LOD 500 - OVERVIEW      a lista de modelos com as áreas em que cada um é
                          cobrado

O que NÃO vem: resultado de auditoria. Os percentuais das planilhas são
agregados de rounds reais, e não dá para reconstruir a resposta item a item a
partir de um agregado — inventá-la encheria a plataforma de dado falso, que é
exatamente o que o usuário recusou. Os modelos entram com a estrutura pronta e
a auditoria em branco.
"""

from __future__ import annotations

import pathlib
import re

import openpyxl
import yaml

BASES = pathlib.Path(r"k:\SPBIM TECH\PLATAFORMAS\Plataforma de auditoria\Bases")
SAIDA = pathlib.Path(__file__).resolve().parent / "dados" / "cpq11-real.yaml"

# Macrodisciplina de cada código de disciplina. O enum tem quatro letras e é
# assim que a coordenação agrupa: A arquitetura, C civil/estrutura, M
# instalações, S sistemas e automação.
MACRO = {
    "ARCH": "A",
    "BASM": "C", "CIVL": "C", "LAND": "C", "STRC": "C",
    "ELEC": "M", "FALM": "M", "FPRT": "M", "MECH": "M", "PLMB": "M",
    "DALI": "S", "DASM": "S", "EPMS": "S", "ROIP": "S", "SCTY": "S",
    "SYST": "S", "TCOM": "S",
}

# Como o nome do arquivo real se decompõe: CPQ11-ARCH-HMET-DATA-R24.rvt
NOME = re.compile(r"^CPQ11-([A-Z0-9]{4})-([A-Z0-9]{4})-([A-Z0-9]{2,4})-(R\d{2})\.(rvt|ifc)$", re.I)


def texto(v) -> str:
    return str(v).strip() if v not in (None, "") else ""


wb500 = openpyxl.load_workbook(BASES / "LOD500_SPECIFIC AUDIT_CONTROL.xlsx", data_only=True)
wb400 = openpyxl.load_workbook(BASES / "LOD400_SPECIFIC AUDIT_CONTROL.xlsx", data_only=True)

# ---------------------------------------------------------------- empresas
base = wb500["BASE"]
empresas: dict[str, str] = {}
for linha in range(3, base.max_row + 1):
    nome, codigo = texto(base.cell(linha, 11).value), texto(base.cell(linha, 12).value)
    if nome and nome not in ("N/A",):
        empresas[nome] = codigo

areas_validas = {
    texto(base.cell(linha, 3).value)
    for linha in range(3, base.max_row + 1)
    if texto(base.cell(linha, 3).value)
}

# ------------------------------------------------- modelos, das abas de área
# Cada aba de área do LOD 400 e do LOD 500 lista os modelos COBRADOS naquela
# área — é assim que a área do escopo de cada disciplina se descobre.
modelos: dict[str, dict] = {}
areas_por_disc: dict[str, set[str]] = {}

for wb, prefixo in ((wb400, "LOD 400 - "), (wb500, "LOD 500 - ")):
    for aba_nome in wb.sheetnames:
        if not aba_nome.startswith(prefixo):
            continue
        area = aba_nome[len(prefixo):].strip()
        if area in ("CONTROL", "OVERVIEW") or not area:
            continue
        aba = wb[aba_nome]
        for linha in range(3, aba.max_row + 1):
            # MAIÚSCULA na origem. O importador normaliza o código, e as
            # planilhas trazem o mesmo arquivo ora como `.rvt`, ora `.RVT` —
            # sem isto os dois viram duas linhas que colidem no UNIQUE.
            arquivo = texto(aba.cell(linha, 2).value).upper()
            m = NOME.match(arquivo)
            if not m:
                continue
            disc, sub = m.group(1).upper(), m.group(2).upper()
            if disc not in MACRO:
                continue
            chave = f"{disc}-{sub}"
            areas_por_disc.setdefault(chave, set()).add(area)
            existente = modelos.get(arquivo)
            if existente is None:
                modelos[arquivo] = {
                    "codigo": arquivo,
                    "disciplina": chave,
                    "instaladora": texto(aba.cell(linha, 5).value) or None,
                    "modeladora": texto(aba.cell(linha, 7).value) or None,
                }

# Só empresas de verdade viram cadastro; o resto vira None no modelo.
for m in modelos.values():
    for papel in ("instaladora", "modeladora"):
        if m[papel] not in empresas:
            m[papel] = None

# ------------------------------------------------------------- disciplinas
disciplinas = []
for chave in sorted(areas_por_disc):
    disc, sub = chave.split("-")
    # Quem projeta: a instaladora mais frequente entre os modelos da disciplina.
    contagem: dict[str, int] = {}
    for m in modelos.values():
        if m["disciplina"] == chave and m["instaladora"]:
            contagem[m["instaladora"]] = contagem.get(m["instaladora"], 0) + 1
    projetista = max(contagem, key=lambda nome: contagem[nome]) if contagem else None

    disciplinas.append(
        {
            "macro": MACRO[disc],
            "disc": disc,
            "sub": sub,
            **({"projetista": projetista} if projetista else {}),
            # Geral em todas — são os 17 itens que todo modelo responde. Os de
            # LOD entram nas que a coordenação de fato audita por área.
            "checklists": ["geral", "lod400", "lod500"],
            "areas": sorted(areas_por_disc[chave]),
        }
    )

doc = {
    "organizacao": {"nome": "SPBIM", "slug": "spbim"},
    "projeto": {
        "codigo": "CPQ11",
        "nome": "CPQ11 — Data Center",
        "cliente": "Microsoft",
        "coordenacao": "SPBIM",
        "bep_ref": "A5.3.2 · Construction BEP",
        "status": "ativo",
    },
    # O padrão REAL dos arquivos do projeto: CPQ11-DISC-SUB-SETOR-SW. Note que
    # não há segmento de MACRO — o exemplo do repositório tinha, e os arquivos
    # entregues não têm. Quem manda é o arquivo.
    "nomenclatura": [
        {"k": "PROJETO", "vals": ["CPQ11"]},
        {"k": "DISC", "vals": sorted({d["disc"] for d in disciplinas})},
        {"k": "SUB", "vals": sorted({d["sub"] for d in disciplinas})},
        {"k": "SETOR", "vals": sorted(areas_validas)},
        {"k": "SW", "vals": ["R22", "R24"]},
    ],
    "empresas": [
        {
            "nome": nome,
            "tipo": "terceirizada",
            "papeis": ["trade", "bim"],
            **({"codigo_acc": codigo} if codigo and codigo != "0000" else {}),
        }
        for nome, codigo in sorted(empresas.items())
    ],
    "disciplinas": disciplinas,
    "modelos": [modelos[k] for k in sorted(modelos)],
}

SAIDA.write_text(
    "# GERADO das planilhas de controle reais de `Bases/` — ver o gerador em\n"
    "# scratchpad/gerar_cpq11.py. Empresas, disciplinas, areas e nomes de\n"
    "# arquivo sao os do projeto; resultado de auditoria NAO vem daqui.\n"
    "#\n"
    "#   python -m scripts.onboarding scripts/dados/cpq11-real.yaml\n\n"
    + yaml.safe_dump(doc, allow_unicode=True, sort_keys=False, width=100),
    encoding="utf-8",
)

print(f"{SAIDA.name}")
print(f"  empresas    {len(doc['empresas'])}")
print(f"  disciplinas {len(disciplinas)}")
print(f"  modelos     {len(modelos)}")
print(f"  areas       {sorted(areas_validas)}")
