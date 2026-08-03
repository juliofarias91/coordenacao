"""Importação de planilha — a ponte provisória da migration 0012.

O QUE ESTE TESTE PROTEGE não é o caminho feliz, é o que as planilhas reais
ensinaram (ver `services/importacao_planilha.py`):

- as abas de LOD 300 têm SEIS layouts de coluna diferentes, então nada pode ser
  lido por índice fixo;
- a porcentagem escrita na planilha está errada em pelo menos duas, então a
  aprovação tem de ser RECONTADA;
- a disciplina da célula está errada em duas, então ela sai do nome do arquivo.

As planilhas são geradas aqui com `openpyxl`, e não lidas do disco: os arquivos
reais estão numa pasta de rede fora do repositório, e um teste que depende dela
falha em qualquer outra máquina. As fábricas reproduzem os DOIS layouts que
importam — o com coluna `INFORMATION` e o sem ela.
"""

from __future__ import annotations

import io

import pytest
from fastapi.testclient import TestClient

from app.services import importacao_planilha
from tests.conftest import BANCO

#: SÓ AS ROTAS PRECISAM DO BANCO. Os testes do leitor são função pura sobre
#: bytes — e são justamente os que protegem as armadilhas das planilhas reais.
#: Deixá-los rodar sem infraestrutura é o que permite conferir o parser numa
#: máquina qualquer, que é onde ele vai ser mexido às pressas.
precisa_banco = pytest.mark.skipif(not BANCO, reason="Postgres indisponível")


def _wb():
    import openpyxl

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    return wb


def _bytes(wb) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def planilha_geral(veredictos: list[str], declarada: float | None = None) -> bytes:
    """A aba `BASE GERAL`: cabeçalho na linha 5, modelo na 3, e a porcentagem
    declarada ACIMA do rótulo — na linha 2, não abaixo dele como no LOD 300."""
    wb = _wb()
    ws = wb.create_sheet("BASE GERAL")
    ws.cell(2, 1, "FILE NAME")
    if declarada is not None:
        ws.cell(2, 7, declarada)
    ws.cell(3, 1, "CPQ04-ARCH-R26")
    # A célula de disciplina MENTE de propósito: é o caso real do arquivo de
    # MECH, que traz "FPRT-FPRT-DATA". Quem manda é o nome do arquivo.
    ws.cell(4, 1, "FPRT-FPRT-DATA")
    for c, rot in enumerate(
        ["INFORMATION", "VERIFICATION", "COMENTARY", "IMAGE", "DIRECTION",
         "ITEMS ANALYZED", "APPROVED (%)"], start=1
    ):
        ws.cell(5, c, rot)
    for i, v in enumerate(veredictos):
        ws.cell(6 + i, 1, f"ITEM {i}")
        ws.cell(6 + i, 2, v)
        ws.cell(6 + i, 3, "-" if v == "APPROVED" else "achado")
        ws.cell(6 + i, 5, "-" if v == "APPROVED" else "corrija")
    return _bytes(wb)


def planilha_lod300(aba: str, veredictos: list[bool], com_information: bool) -> bytes:
    """A aba de LOD 300, nos dois layouts que os arquivos reais têm.

    `com_information=False` reproduz a planilha de ARCH, que NÃO tem a coluna
    `INFORMATION` — lá o nome do item mora em `REVIT PARAMETER`. É o caso que
    quebra qualquer leitura por índice fixo.
    """
    wb = _wb()
    ws = wb.create_sheet(aba)
    cabecalho = ["IMAGE", "ELEMENT", "BIM FORUM REFERENCE", "LOD"]
    if com_information:
        cabecalho.append("INFORMATION")
    cabecalho += ["DATA TYPE", "BIM FORUM DESCRIPTION", "REVIT PARAMETER",
                  "PARAMETER", "VERIFICATION", "COMMENTS"]
    # Deslocado uma coluna à direita, como nos arquivos reais.
    for c, rot in enumerate(cabecalho, start=2):
        ws.cell(5, c, rot)
    cols = {rot: c for c, rot in enumerate(cabecalho, start=2)}

    ws.cell(3, 2, "FILE NAME")
    ws.cell(4, 2, "CPQ4-ELEC-R26.rvt")
    ws.cell(3, 6, "ANALYSED VERSION\n(V0 - DD/MM/AAAA)")
    ws.cell(4, 6, "V4 - 25/06/2026")

    c_nome = cols["INFORMATION"] if com_information else cols["REVIT PARAMETER"]
    for i, ok in enumerate(veredictos):
        linha = 6 + i
        # O grupo só na PRIMEIRA linha do bloco — é como a mesclagem chega.
        if i % 3 == 0:
            ws.cell(linha, cols["ELEMENT"], f"GRUPO {i // 3}")
        ws.cell(linha, c_nome, f"Param {i}")
        ws.cell(linha, cols["VERIFICATION"], ok)
        if not ok:
            ws.cell(linha, cols["COMMENTS"], "Parâmetro sem preenchimento")
    return _bytes(wb)


# ------------------------------------------------------------------ o parser


def test_geral_reconta_em_vez_de_confiar_na_planilha() -> None:
    """A prova de fogo: a planilha DECLARA 0,30 e a conta certa é 0,60.

    Não é hipótese — é a aba STRC dos arquivos reais, cuja fórmula é
    `=COUNTIF(I6:I33, TRUE)/COUNTA(I6:I65)`: alguém acrescentou linhas e só
    arrastou o denominador. Se algum dia isto passar a mostrar 0,30, o
    dashboard voltou a repetir o erro da planilha em vez de corrigi-lo.
    """
    conteudo = planilha_geral(["APPROVED"] * 6 + ["NOT APPROVED"] * 4, declarada=0.30)
    (lida,) = importacao_planilha.ler("DANTE 2 _ STRC _ AUDITORIA GERAL .xlsx", conteudo)

    assert lida.aprovacao == pytest.approx(0.6)
    assert lida.aprovacao_declarada == pytest.approx(0.30)
    assert lida.aprovados == 6
    assert len(lida.itens) == 10


def test_disciplina_sai_do_nome_do_arquivo_e_nao_da_celula() -> None:
    """A célula diz FPRT-FPRT-DATA; o arquivo diz MECH. Vence o arquivo."""
    conteudo = planilha_geral(["APPROVED"] * 3)
    (lida,) = importacao_planilha.ler("DANTE 2 _ MECH _ AUDITORIA GERAL .xlsx", conteudo)
    assert lida.disciplina == "MECH"


def test_not_approved_nao_conta_como_approved() -> None:
    """"NOT APPROVED" CONTÉM "APPROVED". Uma comparação por substring na ordem
    errada aprova a planilha inteira — e o número sairia bonito e falso."""
    conteudo = planilha_geral(["NOT APPROVED"] * 5)
    (lida,) = importacao_planilha.ler("x _ ARCH _ y.xlsx", conteudo)
    assert lida.aprovados == 0
    assert lida.aprovacao == 0.0


@pytest.mark.parametrize("com_information", [True, False])
def test_lod300_le_os_dois_layouts_de_coluna(com_information: bool) -> None:
    """Sem a coluna `INFORMATION`, o nome do item está em `REVIT PARAMETER` —
    é a planilha de ARCH. Leitura por índice fixo devolveria zero item nela."""
    conteudo = planilha_lod300("ELEC", [True, True, False, False], com_information)
    (lida,) = importacao_planilha.ler("DANTE 2 _ ELEC _ x.xlsx", conteudo)

    assert lida.tipo == "lod300"
    assert len(lida.itens) == 4
    assert lida.aprovacao == pytest.approx(0.5)
    assert lida.modelo == "CPQ4-ELEC-R26.rvt"
    assert lida.versao == "V4 - 25/06/2026"


def test_lod300_carrega_o_grupo_mesclado_para_baixo() -> None:
    """O ELEMENT só aparece na primeira linha do bloco; as outras herdam. Sem
    isso, dois terços dos itens ficariam sem grupo no dashboard."""
    conteudo = planilha_lod300("STRC", [True] * 6, com_information=True)
    (lida,) = importacao_planilha.ler("a _ STRC _ b.xlsx", conteudo)
    assert [i.grupo for i in lida.itens] == ["GRUPO 0"] * 3 + ["GRUPO 1"] * 3


def test_um_arquivo_com_as_duas_abas_devolve_as_duas() -> None:
    """Os arquivos da pasta `LOD 300` trazem `BASE GERAL` E a aba da
    disciplina. Importar um deles tem de render as duas auditorias."""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(planilha_geral(["APPROVED"] * 4)))
    outra = openpyxl.load_workbook(
        io.BytesIO(planilha_lod300("ELEC", [True, False], True))
    )
    origem = outra["ELEC"]
    destino = wb.create_sheet("ELEC")
    for linha in origem.iter_rows():
        for celula in linha:
            destino.cell(celula.row, celula.column, celula.value)
    # Uma aba GUIDE no meio: ela existe nos arquivos reais e deve ser ignorada.
    wb.create_sheet("GUIDE LOD300")

    lidas = importacao_planilha.ler("DANTE 2 _ ELEC _ x.xlsx", _bytes(wb))
    assert {p.tipo for p in lidas} == {"geral", "lod300"}


def test_arquivo_sem_aba_conhecida_e_recusado() -> None:
    wb = _wb()
    wb.create_sheet("Planilha1").cell(1, 1, "qualquer coisa")
    with pytest.raises(importacao_planilha.PlanilhaInvalida):
        importacao_planilha.ler("solto.xlsx", _bytes(wb))


# -------------------------------------------------------------------- a rota


def _subir(client: TestClient, arquivos: list[tuple[str, bytes]], projeto_id=None):
    query = f"?projeto_id={projeto_id}" if projeto_id else ""
    return client.post(
        f"/api/v1/importacao/planilhas{query}",
        files=[
            ("arquivos", (nome, dados,
                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))
            for nome, dados in arquivos
        ],
    )


@precisa_banco
def test_upload_multiplo_e_tolerante_a_arquivo_ruim(autenticado: TestClient) -> None:
    """Uma planilha corrompida no lote NÃO pode derrubar as outras. Subir
    catorze e receber 400 por causa da décima obriga a descobrir qual e
    recomeçar — o lote inteiro."""
    wb = _wb()
    wb.create_sheet("Planilha1")

    r = _subir(
        autenticado,
        [
            ("DANTE 2 _ ARCH _ AUDITORIA GERAL .xlsx", planilha_geral(["APPROVED"] * 4)),
            ("lixo.xlsx", _bytes(wb)),
            ("DANTE 2 _ ELEC _ AUDITORIA GERAL .xlsx",
             planilha_geral(["APPROVED", "NOT APPROVED"])),
        ],
    )
    assert r.status_code == 201, r.text
    corpo = r.json()
    assert len(corpo["importadas"]) == 2
    assert len(corpo["recusadas"]) == 1
    assert corpo["recusadas"][0]["arquivo"] == "lixo.xlsx"


@precisa_banco
def test_reimportar_substitui_em_vez_de_duplicar(autenticado: TestClient) -> None:
    """Subir de novo o arquivo da mesma disciplina TROCA o anterior. Sem isso a
    média conta o mesmo modelo duas vezes e anda sozinha a cada upload."""
    nome = "DANTE 2 _ ARCH _ AUDITORIA GERAL .xlsx"
    _subir(autenticado, [(nome, planilha_geral(["APPROVED"] * 4))])
    _subir(autenticado, [(nome, planilha_geral(["NOT APPROVED"] * 4))])

    dados = autenticado.get("/api/v1/importacao/dashboard").json()
    assert dados["total"]["planilhas"] == 1
    assert dados["total"]["aprovacao"] == 0.0


@precisa_banco
def test_dashboard_pondera_pelos_itens(autenticado: TestClient) -> None:
    """A média NÃO é a média das porcentagens.

    Uma planilha de 10 itens com 100% e outra de 90 itens com 0% dão 10% de
    aprovação real — não 50%. Tratar as duas como iguais faria o modelo menos
    auditado pesar o mesmo que o mais auditado.
    """
    _subir(
        autenticado,
        [
            ("a _ ARCH _ x.xlsx", planilha_geral(["APPROVED"] * 10)),
            ("b _ ELEC _ x.xlsx", planilha_geral(["NOT APPROVED"] * 90)),
        ],
    )
    total = autenticado.get("/api/v1/importacao/dashboard").json()["total"]
    assert total["itens"] == 100
    assert total["aprovacao"] == pytest.approx(0.10)


@precisa_banco
def test_dashboard_separa_geral_de_lod300_e_lista_o_que_mais_reprova(
    autenticado: TestClient,
) -> None:
    _subir(
        autenticado,
        [
            ("a _ ARCH _ x.xlsx", planilha_geral(["NOT APPROVED"] * 3)),
            ("b _ ELEC _ x.xlsx", planilha_geral(["NOT APPROVED"] * 3)),
            ("c _ STRC _ x.xlsx", planilha_lod300("STRC", [True, False], True)),
        ],
    )
    dados = autenticado.get("/api/v1/importacao/dashboard").json()

    assert {f["rotulo"] for f in dados["por_tipo"]} == {"geral", "lod300"}
    assert {f["rotulo"] for f in dados["por_disciplina"]} == {"ARCH", "ELEC", "STRC"}
    # ITEM 0..2 reprovam nas DUAS planilhas gerais; o de LOD aparece em uma só e
    # fica de fora, porque um caso isolado não é padrão.
    assert dados["criticos"]
    assert all(c["ocorrencias"] > 1 for c in dados["criticos"])
    assert dados["criticos"][0]["taxa"] == 1.0
